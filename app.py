import streamlit as st
import calendar
import io
import os
import pandas as pd
import bcrypt

from datetime import date, timedelta
from supabase import create_client, Client


SUPABASE_URL = st.secrets.get("SUPABASE_URL")
SUPABASE_KEY = st.secrets.get("SUPABASE_ANON_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    st.error("❌ Supabase non configuré (secrets manquants)")
    st.stop()

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

TRANSPORTS = ["Voiture", "Vélo", "Transport"]
TAUX = {
    "Voiture": 0.10,
    "Vélo": 0.00,
    "Transport": 0.00
}

PLAFOND_MENSUEL = 60.0

JOURS_FR = [
    "Lundi", "Mardi", "Mercredi",
    "Jeudi", "Vendredi", "Samedi", "Dimanche"
]

MOIS_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]
def sb_select(table, filters=None):
    q = supabase.table(table).select("*")
    if filters:
        for k, v in filters.items():
            q = q.eq(k, v)
    return q.execute().data or []


def sb_insert(table, data):
    return supabase.table(table).insert(data).execute()


def sb_update(table, data, filters):
    q = supabase.table(table).update(data)
    for k, v in filters.items():
        q = q.eq(k, v)
    return q.execute()


def sb_delete(table, filters):
    q = supabase.table(table).delete()
    for k, v in filters.items():
        q = q.eq(k, v)
    return q.execute()

def get_periode_reference(annee: int, mois: int):
    dernier_jour = calendar.monthrange(annee, mois)[1]

    start = date(annee, mois, 1)
    end = date(annee, mois, dernier_jour)

    mois_comptable = f"{annee}-{mois:02d}"

    return start, end, mois_comptable



def total_mois_courant(user_id, annee, mois, km):
    """
    Calcule le montant brut et le montant plafonné
    d'un mois civil.
    """

    dernier_jour = calendar.monthrange(annee, mois)[1]

    start = date(annee, mois, 1)
    end = date(annee, mois, dernier_jour)

    rows = (
        supabase
        .table("trajets")
        .select("transport, km_utilise")
        .eq("user_id", user_id)
        .gte("jour", start.isoformat())
        .lte("jour", end.isoformat())
        .execute()
        .data
    )

    if not rows:
        return 0.0, 0.0

    df = pd.DataFrame(rows)

    # Taux selon le moyen de transport
    df["taux"] = df["transport"].map(TAUX).fillna(0)

    # Utilise le km enregistré sur le trajet
    # sinon reprend le km actuel de l'utilisateur
    df["km_effectif"] = df["km_utilise"].fillna(km)

    # Calcul de chaque indemnité
    df["indemnite"] = df["taux"] * df["km_effectif"]

    brut = round(df["indemnite"].sum(), 2)
    plafonne = round(min(brut, PLAFOND_MENSUEL), 2)

    return brut, plafonne


def total_mois_periode(user_id, periode_start, periode_end, km):
    rows = (
        supabase
        .table("trajets")
        .select("transport")
        .eq("user_id", user_id)
        .gte("jour", periode_start.isoformat())
        .lte("jour", periode_end.isoformat())
        .execute()
        .data
    )

    if not rows:
        return 0.0, 0.0

    df = pd.DataFrame(rows)

    # ✅ calcul journalier correct
    df["taux"] = df["transport"].map(TAUX)
    df["km_effectif"] = df["km_utilise"].fillna(km)
    df["indemnite_jour"] = df["taux"] * df["km_effectif"]

    brut = df["indemnite_jour"].sum()
    plafonne = min(brut, PLAFOND_MENSUEL)

    return round(brut, 2), round(plafonne, 2)


def valider_mois(user_id, annee, mois, km):
    """
    Validation d'un mois civil.

    - Première validation :
        -> enregistre le mois.

    - Validation d'un mois déjà validé :
        -> recalcule le montant
        -> crée/met à jour la régularisation
        -> met à jour la validation
    """

    mois_courant = f"{annee}-{mois:02d}"

    # ==========================================
    # Calcul du mois
    # ==========================================
    brut, plafonne = total_mois_courant(
        user_id,
        annee,
        mois,
        km
    )

    # ==========================================
    # Recherche validation existante
    # ==========================================
    existing = (
        supabase
        .table("validations")
        .select("*")
        .eq("user_id", user_id)
        .eq("mois", mois_courant)
        .execute()
        .data
    )

    regularisation = 0.0

    # ==========================================
    # PREMIÈRE VALIDATION
    # ==========================================
    if not existing:

        supabase.table("validations").insert({
            "user_id": user_id,
            "mois": mois_courant,
            "km_utilise": km,
            "brut": brut,
            "plafonne": plafonne
        }).execute()

    # ==========================================
    # MOIS DÉJÀ VALIDÉ
    # ==========================================
    else:

        ancienne_validation = existing[0]

        ancien_plafonne = float(
            ancienne_validation["plafonne"]
        )

        regularisation = creer_regularisation_si_necessaire(
            user_id=user_id,
            mois_source=mois_courant,
            ancien_plafonne=ancien_plafonne,
            nouveau_plafonne=plafonne
        )

        supabase.table("validations") \
            .update({
                "km_utilise": km,
                "brut": brut,
                "plafonne": plafonne
            }) \
            .eq("user_id", user_id) \
            .eq("mois", mois_courant) \
            .execute()

    # ==========================================
    # Total payé
    # ==========================================
    total = round(plafonne + regularisation, 2)

    return {
        "mois": mois_courant,
        "brut": round(brut, 2),
        "plafonne": round(plafonne, 2),
        "regularisation": round(regularisation, 2),
        "total_paye": total
    }



def total_deja_paye(user_id, mois):
    """
    Retourne le TOTAL réellement payé pour un mois donné.
    Inclut :
    - le montant plafonné validé
    - toutes les régularisations provenant de ce mois
    """

    total = 0.0

    # ===== validation du mois =====
    res_val = (
        supabase
        .table("validations")
        .select("plafonne")
        .eq("user_id", user_id)
        .eq("mois", mois)
        .execute()
        .data
    )

    if res_val:
        total += float(res_val[0]["plafonne"])

    # ===== régularisations issues de ce mois =====
    res_reg = (
        supabase
        .table("regularisations")
        .select("montant")
        .eq("user_id", user_id)
        .eq("mois_source", mois)
        .execute()
        .data
    )

    if res_reg:
        total += sum(float(r["montant"]) for r in res_reg)

    return round(total, 2)


def calcul_regularisation_mois_precedent(
    user_id,
    mois_courant,
    montant_mois_courant
):

    an, m = map(int, mois_courant.split("-"))

    # =========================
    # JANVIER = PAS DE MOIS PRÉCÉDENT
    # =========================
    if m == 1:
        return 0.0, None

    mois_prec = f"{an}-{m-1:02d}"

    # =========================
    # TOTAL DÉJÀ PAYÉ
    # (validation + anciennes régularisations)
    # =========================
    deja_paye = total_deja_paye(user_id, mois_prec)

    # =========================
    # SI PLAFOND ATTEINT
    # =========================
    if deja_paye >= PLAFOND_MENSUEL:
        return 0.0, mois_prec

    # =========================
    # MANQUE À COMPLÉTER
    # =========================
    manque = PLAFOND_MENSUEL - deja_paye

    # =========================
    # RÉGULARISATION
    # =========================
    regul = min(montant_mois_courant, manque)

    return round(regul, 2), mois_prec



def total_mois_comptable(user_id, mois_comptable, km):
    """
    Calcule le total réel d'un mois civil
    (du 1er au dernier jour du mois).
    """

    annee, mois = map(int, mois_comptable.split("-"))

    start = date(annee, mois, 1)
    end = date(annee, mois, calendar.monthrange(annee, mois)[1])

    rows = (
        supabase
        .table("trajets")
        .select("transport, km_utilise")
        .eq("user_id", user_id)
        .gte("jour", start.isoformat())
        .lte("jour", end.isoformat())
        .execute()
        .data
    )

    if not rows:
        return 0.0, 0.0

    df = pd.DataFrame(rows)

    df["taux"] = df["transport"].map(TAUX)
    df["km_effectif"] = df["km_utilise"].fillna(km)
    df["indemnite_jour"] = df["taux"] * df["km_effectif"]

    brut = df["indemnite_jour"].sum()
    plafonne = min(brut, PLAFOND_MENSUEL)

    return round(brut, 2), round(plafonne, 2)



def total_regularisations(user_id, mois_cible):
    rows = (
        supabase
        .table("regularisations")
        .select("montant")
        .eq("user_id", user_id)
        .eq("mois_cible", mois_cible)
        .execute()
        .data
    )

    if not rows:
        return 0.0

    return round(sum(r["montant"] for r in rows), 2)


def creer_regularisation_si_necessaire(
    user_id,
    mois_source,
    ancien_plafonne,
    nouveau_plafonne
):
    """
    Si un mois déjà validé est modifié,
    crée (ou met à jour) UNE SEULE régularisation
    vers le mois suivant.
    """

    ancien_plafonne = round(float(ancien_plafonne), 2)
    nouveau_plafonne = round(float(nouveau_plafonne), 2)

    diff = round(nouveau_plafonne - ancien_plafonne, 2)

    # Aucun changement
    if diff == 0:
        return 0.0

    # -----------------------------
    # Mois cible = mois suivant
    # -----------------------------
    annee, mois = map(int, mois_source.split("-"))

    if mois == 12:
        mois_cible = f"{annee+1}-01"
    else:
        mois_cible = f"{annee}-{mois+1:02d}"

    # -----------------------------
    # Suppression ancienne régul
    # -----------------------------
    supabase.table("regularisations") \
        .delete() \
        .eq("user_id", user_id) \
        .eq("mois_source", mois_source) \
        .execute()

    # -----------------------------
    # Création nouvelle régul
    # -----------------------------
    supabase.table("regularisations").insert({
        "user_id": user_id,
        "mois_source": mois_source,
        "mois_cible": mois_cible,
        "montant": diff
    }).execute()

    return diff



def montant_final_paye(user_id, periode_start, periode_end, mois_comptable, km):
    brut, plafonne = total_mois_periode(
        user_id,
        periode_start,
        periode_end,
        km
    )

    regul = total_regularisations(user_id, mois_comptable)

    total = round(plafonne + regul, 2)

    return {
        "brut": round(brut, 2),
        "plafonne": round(plafonne, 2),
        "regularisation": round(regul, 2),
        "total_paye": total,
        "plafond_atteint": brut > PLAFOND_MENSUEL
    }

def login_user(login, pwd):
    try:
        res = (
            supabase
            .table("users")
            .select("*")
            .eq("login", login)
            .execute()
        )

        if not res.data:
            return None

        user = res.data[0]

        if bcrypt.checkpw(
            pwd.encode(),
            user["password"].encode()
        ):
            return user

        return None

    except Exception as e:
        st.error(f"Erreur login Supabase : {e}")
        return None


def init_admin():
    res = (
        supabase
        .table("users")
        .select("id")
        .eq("login", "admin")
        .execute()
    )

    if res.data:
        return  # admin déjà présent

    hpwd = bcrypt.hashpw(b"admin", bcrypt.gensalt()).decode()

    supabase.table("users").insert({
        "nom": "Admin",
        "prenom": "Root",
        "login": "admin",
        "password": hpwd,
        "km": 0,
        "is_admin": True,
        "must_change_pwd": True
    }).execute()


# ⚠️ À appeler UNE SEULE FOIS
init_admin()


if "user" not in st.session_state:
    st.session_state.user = None

if "edit_mode" not in st.session_state:
    st.session_state.edit_mode = False

if "jours_selectionnes" not in st.session_state:
    st.session_state.jours_selectionnes = set()


if st.session_state.user is None:
    st.set_page_config("Indemnités kilométriques", layout="wide")
    st.title("Connexion")

    login_input = st.text_input("Login")
    pwd = st.text_input("Mot de passe", type="password")

    if st.button("Connexion"):
        u = login_user(login_input, pwd)
        if u:
            st.session_state.user = u
            st.rerun()
        else:
            st.error("Identifiants incorrects")

    st.stop()


uid = st.session_state.user["id"]
nom = st.session_state.user["nom"]
prenom = st.session_state.user["prenom"]
login = st.session_state.user["login"]
km = st.session_state.user["km"]
is_admin = st.session_state.user["is_admin"]
must_change_pwd = st.session_state.user["must_change_pwd"]

st.set_page_config("Indemnités kilométriques", layout="wide")

st.markdown("""
<style>
.day-box {
    padding: 4px;
    border-radius: 6px;
}
.day-locked {
    opacity: 0.45;
}
.day-new {
    color: #2ecc71;
    font-size: 0.8em;
}
.day-info {
    font-size: 0.75em;
}
</style>
""", unsafe_allow_html=True)


if must_change_pwd:
    st.warning("Vous devez changer votre mot de passe")

    p1 = st.text_input("Nouveau mot de passe", type="password")
    p2 = st.text_input("Confirmation", type="password")

    if st.button("Changer le mot de passe"):
        if not p1 or p1 != p2:
            st.error("Mot de passe invalide")
        else:
            hpwd = bcrypt.hashpw(p1.encode(), bcrypt.gensalt()).decode()

            supabase.table("users").update({
                "password": hpwd,
                "must_change_pwd": False
            }).eq("id", uid).execute()

            st.success("Mot de passe modifié")
            st.session_state.user = None
            st.rerun()

    st.stop()

st.sidebar.write(f"👤 {prenom} {nom}")

if st.sidebar.button("Déconnexion", key="logout_btn"):
    st.session_state.user = None
    st.rerun()

menu = st.sidebar.radio(
    "Menu",
    ["Encodage", "Historique"] if not is_admin else
    ["Utilisateurs", "Encodage", "Validation", "Exports"],
    key="menu_radio"
)

annee = mois_num = None
periode_start = periode_end = None
mois_comptable = None

if menu in ["Encodage", "Validation", "Exports"]:

    today = date.today()

    annee = st.selectbox(
        "Année",
        range(today.year - 1, today.year + 3),
        index=1,
        key="select_annee"
    )

    mois_label = st.selectbox(
        "Mois",
        MOIS_FR,
        index=today.month - 1,
        key="select_mois"
    )

    mois_num = MOIS_FR.index(mois_label) + 1

    periode_start, periode_end, mois_comptable = get_periode_reference(
        annee,
        mois_num
    )

    st.caption(
        f"📅 Période de référence : "
        f"{periode_start.strftime('%d/%m/%Y')} → "
        f"{periode_end.strftime('%d/%m/%Y')}"
    )

def calendrier(user_id, admin=False):

    # ===== récupération trajets période =====
    rows = (
        supabase
        .table("trajets")
        .select("jour, transport, validated, sent_for_validation")
        .eq("user_id", user_id)
        .gte("jour", periode_start.isoformat())
        .lte("jour", periode_end.isoformat())
        .execute()
        .data
    )

    data = {
        date.fromisoformat(r["jour"]): {
            "transport": r["transport"],
            "validated": bool(r["validated"]),
            "sent": bool(r.get("sent_for_validation", False))
        }
        for r in rows
    }

    st.markdown(f"""
**Période affichée**  
📅 {periode_start.strftime('%d/%m/%Y')} → {periode_end.strftime('%d/%m/%Y')}
""")

    jours = []

    # ===== génération des jours de la période =====
    current = periode_start
    days = []

    while current <= periode_end:
        days.append(current)
        current += timedelta(days=1)

    # ===== alignement semaine (lundi → dimanche) =====
    first_weekday = days[0].weekday()
    for _ in range(first_weekday):
        days.insert(0, None)

    # ===== affichage par semaines =====
    for i in range(0, len(days), 7):

        cols = st.columns(7)
        week = days[i:i+7]

        for col, day in zip(cols, week):

            with col:

                if day is None:
                    st.write("")
                    continue

                wd = day.weekday()
                is_weekend = wd >= 5

                existe = day in data
                validated = existe and data[day]["validated"]
                sent = existe and data[day]["sent"]

                locked = (
                    is_weekend
                    or ((sent or validated) and not admin)
                )

                label = f"{JOURS_FR[wd][:2]} {day.strftime('%d/%m')}"

                # ==========================
                # WEEK-END
                # ==========================
                if is_weekend:

                    st.checkbox(label, disabled=True)
                    st.caption("⛔ Week-end")

                    val = False

                # ==========================
                # JOUR EXISTANT
                # ==========================
                elif existe:

                    icon = {
                        "Voiture": "🚗",
                        "Vélo": "🚲",
                        "Transport": "🚌"
                    }.get(data[day]["transport"], "❓")

                    status = "✅" if validated else "⏳" if sent else ""

                    # ===== MODE ÉDITION =====
                    if st.session_state.edit_mode:

                        selected = day.day in st.session_state.jours_selectionnes

                        toggle = st.checkbox(
                            label,
                            value=selected,
                            disabled=locked
                        )

                        if toggle:
                            st.session_state.jours_selectionnes.add(day.day)
                        else:
                            st.session_state.jours_selectionnes.discard(day.day)

                        st.caption(f"{icon} {data[day]['transport']} {status}")

                        val = True

                    # ===== MODE NORMAL =====
                    else:

                        st.checkbox(label, value=True, disabled=True)
                        st.caption(f"{icon} {data[day]['transport']} {status}")

                        val = True

                # ==========================
                # NOUVEAU JOUR
                # ==========================
                else:

                    val = st.checkbox(label, disabled=is_weekend)

                    if val:
                        st.caption("➕ Nouveau")

                if not is_weekend:
                    jours.append((day, val, existe, validated, sent))

    return jours

save_clicked = False
send_clicked = False




if menu == "Encodage":

    st.header(
        f"Encodage des trajets — "
        f"{periode_start.strftime('%d/%m/%Y')} → {periode_end.strftime('%d/%m/%Y')}"
    )
    st.divider()

    # ==================================================
    # INIT STATE (sécurisé)
    # ==================================================
    if "edit_mode" not in st.session_state:
        st.session_state.edit_mode = False

    if "jours_selectionnes" not in st.session_state:
        st.session_state.jours_selectionnes = set()

    # ==================================================
    # UTILISATEUR CIBLE
    # ==================================================
    if is_admin:

        res_users = (
            supabase
            .table("users")
            .select("id, prenom, nom, km")
            .neq("login", "admin")
            .order("nom")
            .execute()
        )

        users = res_users.data or []

        if not users:
            st.warning("Aucun utilisateur disponible.")
            st.stop()

        labels = []
        user_map = {}

        for u in users:
            label = f"{u['prenom']} {u['nom']}"
            labels.append(label)
            user_map[label] = u["id"]


        
        selected_label = st.selectbox("Utilisateur à encoder", labels)

        selected_user = next(
            u for u in users
            if f"{u['prenom']} {u['nom']}" == selected_label
        )

        cible = selected_user["id"]
        km_cible = float(selected_user["km"] or 0)




    else:
        cible = uid
        km_cible = float(km)


    # ==================================================
    # TRANSPORT PAR DÉFAUT
    # ==================================================
    transport_global = st.selectbox(
        "Moyen de transport pour les nouveaux jours",
        TRANSPORTS
    )

    changer_transport = st.checkbox(
        "Changer le transport des jours existants sélectionnés",
        value=False
    )

    st.divider()

    # ==================================================
    # ACTIONS
    # ==================================================
    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("✏️ Modifier l’encodage"):
            st.session_state.edit_mode = True
            st.session_state.jours_selectionnes = set()

    with col2:
        save_clicked = st.button("💾 Enregistrer")

    with col3:
        delete_selected = st.button("🗑 Supprimer sélection")

    if st.session_state.edit_mode:
        st.info(
            "Mode modification actif : sélectionnez les jours existants "
            "puis cliquez sur Enregistrer ou Supprimer."
        )

    # ==================================================
    # CALENDRIER
    # ==================================================
    jours = calendrier(cible, admin=is_admin)

    # ==================================================
    # 🟢 SUPPRESSION (CORRIGÉE + SORTIE MODE)
    # ==================================================
    if delete_selected and st.session_state.edit_mode:

        jours_supprimes = 0

        for day, val, existe, validated, sent in jours:

            if existe and day.day in st.session_state.jours_selectionnes:

                supabase.table("trajets") \
                    .delete() \
                    .eq("user_id", cible) \
                    .eq("jour", day.isoformat()) \
                    .execute()

                jours_supprimes += 1

        # RESET COMPLET
        st.session_state.jours_selectionnes = set()
        st.session_state.edit_mode = False   # 🔥 SORTIE DU MODE

        st.success(f"{jours_supprimes} jour(s) supprimé(s).")
        st.rerun()

    # ==================================================
    # ENREGISTREMENT
    # ==================================================
    def enregistrer_encodage():
        """Enregistre les ajouts / suppressions / modifications
        de la sélection courante du calendrier."""

        for day, val, existe, validated, sent in jours:

            jour_iso = day.isoformat()

            if not (periode_start <= day <= periode_end):
                continue

            # AJOUT
            if val and not existe:
                supabase.table("trajets").insert({
                    "user_id": cible,
                    "jour": jour_iso,
                    "transport": transport_global,
                    "km_utilise": km_cible,
                    "validated": False,
                    "sent_for_validation": False
                }).execute()

            # SUPPRESSION via décochage
            elif existe and st.session_state.edit_mode and not val:

                supabase.table("trajets") \
                    .delete() \
                    .eq("user_id", cible) \
                    .eq("jour", jour_iso) \
                    .execute()

            # 🟢 MODIFICATION CORRIGÉE
            elif (
                existe
                and st.session_state.edit_mode
                and changer_transport
                and day.day in st.session_state.jours_selectionnes
            ):

                supabase.table("trajets") \
                    .update({
                        "transport": transport_global,
                        "validated": False
                    }) \
                    .eq("user_id", cible) \
                    .eq("jour", jour_iso) \
                    .execute()

    if save_clicked:

        enregistrer_encodage()

        # RESET
        st.session_state.edit_mode = False
        st.session_state.jours_selectionnes = set()

        st.success("Encodage enregistré.")
        st.rerun()

    # ==================================================
    # ENVOI POUR VALIDATION
    # ==================================================
    if not is_admin and not st.session_state.edit_mode:

        if st.button("📤 Envoyer pour validation"):

            # Enregistre d'abord les jours cochés mais pas encore
            # sauvegardés, pour que l'envoi fonctionne en un seul clic.
            enregistrer_encodage()

            supabase.table("trajets") \
                .update({"sent_for_validation": True}) \
                .eq("user_id", cible) \
                .gte("jour", periode_start.isoformat()) \
                .lte("jour", periode_end.isoformat()) \
                .execute()

            st.session_state.edit_mode = False
            st.session_state.jours_selectionnes = set()

            st.success("Trajets enregistrés et période envoyée pour validation.")
            st.rerun()



if menu == "Utilisateurs":

    if not is_admin:
        st.error("Accès refusé")
        st.stop()

    st.header("Gestion des utilisateurs")
    st.divider()

    # ==================================================
    # RÉCUPÉRATION USERS
    # ==================================================
    res_users = (
        supabase
        .table("users")
        .select("id, nom, prenom, login, km, is_admin")
        .order("nom")
        .execute()
    )

    users = res_users.data or []

    # ==================================================
    # CRÉATION UTILISATEUR
    # ==================================================
    st.subheader("➕ Créer un utilisateur")

    with st.form("form_create_user"):
        nom_u = st.text_input("Nom")
        prenom_u = st.text_input("Prénom")
        login_u = st.text_input("Login")
        km_u = st.number_input("Kilomètres domicile-travail", min_value=0.0)
        pwd_u = st.text_input("Mot de passe", type="password")
        is_admin_u = st.checkbox("Administrateur")

        submit_create = st.form_submit_button("Créer")

        if submit_create:
            if not login_u or not pwd_u:
                st.error("Login et mot de passe obligatoires.")
            else:
                hpwd = bcrypt.hashpw(
                    pwd_u.encode(),
                    bcrypt.gensalt()
                ).decode()

                supabase.table("users").insert({
                    "nom": nom_u.strip(),
                    "prenom": prenom_u.strip(),
                    "login": login_u.lower().strip(),
                    "password": hpwd,
                    "km": float(km_u),
                    "is_admin": bool(is_admin_u),
                    "must_change_pwd": True
                }).execute()

                st.success("Utilisateur créé")
                st.rerun()

    st.divider()

    # ==================================================
    # MODIFICATION / SUPPRESSION
    # ==================================================
    st.subheader("✏️ Modifier / supprimer")

    if not users:
        st.info("Aucun utilisateur")
        st.stop()

    user_labels = {
        f"{u['prenom']} {u['nom']} ({u['login']})": u
        for u in users
    }

    selected_label = st.selectbox(
        "Utilisateur",
        list(user_labels.keys())
    )

    user = user_labels[selected_label]

    with st.form("form_edit_user"):

        nom_e = st.text_input("Nom", value=user["nom"])
        prenom_e = st.text_input("Prénom", value=user["prenom"])
        login_e = st.text_input("Login", value=user["login"])

        km_e = st.number_input(
            "KM",
            min_value=0.0,
            value=float(user["km"])
        )

        is_admin_e = st.checkbox(
            "Administrateur",
            value=user["is_admin"]
        )

        st.markdown("### 🔐 Mot de passe")

        new_pwd = st.text_input("Nouveau mot de passe", type="password")
        reset_pwd = st.checkbox("Réinitialiser à '1234'")

        col1, col2 = st.columns(2)
        save = col1.form_submit_button("💾 Enregistrer")
        delete = col2.form_submit_button("🗑 Supprimer")

        if save:

            data = {
                "nom": nom_e.strip(),
                "prenom": prenom_e.strip(),
                "login": login_e.lower().strip(),
                "km": float(km_e),
                "is_admin": bool(is_admin_e)
            }

            if reset_pwd:
                hpwd = bcrypt.hashpw("1234".encode(), bcrypt.gensalt()).decode()
                data["password"] = hpwd
                data["must_change_pwd"] = True

            elif new_pwd:
                hpwd = bcrypt.hashpw(new_pwd.encode(), bcrypt.gensalt()).decode()
                data["password"] = hpwd
                data["must_change_pwd"] = True

            supabase.table("users") \
                .update(data) \
                .eq("id", user["id"]) \
                .execute()

            st.success("Utilisateur modifié")
            st.rerun()

        if delete:

            if user["id"] == uid:
                st.error("Impossible de supprimer votre propre compte.")
                st.stop()

            supabase.table("users") \
                .delete() \
                .eq("id", user["id"]) \
                .execute()

            st.warning("Utilisateur supprimé")
            st.rerun()

    st.divider()

    # ==================================================
    # IMPORT UTILISATEURS
    # ==================================================
    st.subheader("📥 Import Excel")

    modele = pd.DataFrame(
        columns=["nom", "prenom", "login", "km", "password", "is_admin"]
    )

    buf_tpl = io.BytesIO()
    modele.to_excel(buf_tpl, index=False)
    buf_tpl.seek(0)

    st.download_button(
        "📄 Télécharger modèle",
        buf_tpl,
        file_name="modele_utilisateurs.xlsx"
    )

    with st.form("form_import_users"):

        file = st.file_uploader("Fichier Excel", type="xlsx")

        submit_import = st.form_submit_button("Importer")

        if submit_import:
            if not file:
                st.warning("Choisissez un fichier")
            else:
                df = pd.read_excel(file)

                colonnes = {
                    "nom", "prenom", "login",
                    "km", "password", "is_admin"
                }

                if not colonnes.issubset(df.columns):
                    st.error("Colonnes invalides")
                else:
                    ok, ko = 0, 0

                    for _, r in df.iterrows():
                        try:
                            hpwd = bcrypt.hashpw(
                                str(r["password"]).encode(),
                                bcrypt.gensalt()
                            ).decode()

                            supabase.table("users").insert({
                                "nom": str(r["nom"]).strip(),
                                "prenom": str(r["prenom"]).strip(),
                                "login": str(r["login"]).lower().strip(),
                                "password": hpwd,
                                "km": float(r["km"]),
                                "is_admin": bool(r["is_admin"]),
                                "must_change_pwd": True
                            }).execute()

                            ok += 1
                        except:
                            ko += 1

                    st.success(f"{ok} ajoutés / {ko} erreurs")
                    st.rerun()

    st.divider()

    # ==================================================
    # EXPORT UTILISATEURS
    # ==================================================
    st.subheader("📊 Export")

    if users:

        df_export = pd.DataFrame(users).rename(columns={
            "nom": "Nom",
            "prenom": "Prénom",
            "login": "Login",
            "km": "KM",
            "is_admin": "Admin"
        })

        st.dataframe(df_export, width="stretch")

        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False)

        buffer.seek(0)

        st.download_button(
            "📥 Télécharger",
            buffer,
            file_name="utilisateurs.xlsx"
        )




def afficher_exports():

    if not is_admin:
        st.error("Accès réservé aux administrateurs.")
        st.stop()

    st.header("Export des indemnités")
    st.caption(
        f"Mois précédent : {MOIS_FR[(mois_num - 2) % 12]} "
        f"{annee - 1 if mois_num == 1 else annee} — "
        f"Mois en cours : {MOIS_FR[mois_num - 1]} {annee}"
    )

    mois_courant = f"{annee}-{mois_num:02d}"
    if mois_num == 1:
        annee_prec, mois_prec_num = annee - 1, 12
    else:
        annee_prec, mois_prec_num = annee, mois_num - 1
    mois_precedent = f"{annee_prec}-{mois_prec_num:02d}"

    debut_prec = date(annee_prec, mois_prec_num, 1)
    fin_courant = date(annee, mois_num, calendar.monthrange(annee, mois_num)[1])

    users = (
        supabase.table("users")
        .select("id, nom, prenom, km")
        .neq("login", "admin")
        .order("nom")
        .execute().data or []
    )

    trajets = (
        supabase.table("trajets")
        .select("user_id, jour, transport, km_utilise")
        .gte("jour", debut_prec.isoformat())
        .lte("jour", fin_courant.isoformat())
        .execute().data or []
    )

    # Une régularisation signale que le total final du mois source a été modifié.
    regularisations = (
        supabase.table("regularisations")
        .select("user_id, mois_source, montant")
        .in_("mois_source", [mois_precedent, mois_courant])
        .execute().data or []
    )
    mois_modifies = {
        (int(r["user_id"]), r["mois_source"])
        for r in regularisations
        if round(float(r["montant"]), 2) != 0
    }

    details = {}
    for trajet in trajets:
        jour = date.fromisoformat(trajet["jour"])
        mois = f"{jour.year}-{jour.month:02d}"
        if mois not in (mois_precedent, mois_courant):
            continue
        cle = (int(trajet["user_id"]), mois)
        details.setdefault(cle, []).append(trajet)

    lignes = []
    for user in users:
        user_id = int(user["id"])
        km_defaut = float(user.get("km") or 0)
        ligne = {
            "Utilisateur": f"{user['prenom']} {user['nom']}",
            "_modifie_prec": (user_id, mois_precedent) in mois_modifies,
            "_modifie_courant": (user_id, mois_courant) in mois_modifies,
        }
        a_des_trajets = False

        for mois, libelle in (
            (mois_precedent, f"{MOIS_FR[mois_prec_num - 1]} {annee_prec}"),
            (mois_courant, f"{MOIS_FR[mois_num - 1]} {annee}"),
        ):
            jours_mois = details.get((user_id, mois), [])
            a_des_trajets = a_des_trajets or bool(jours_mois)
            montants = []

            for transport in TRANSPORTS:
                jours_transport = [
                    t for t in jours_mois if t.get("transport") == transport
                ]
                ligne[f"{transport} — {libelle}"] = len(jours_transport)
                montants.extend(
                    TAUX.get(transport, 0) * float(
                        t.get("km_utilise") if t.get("km_utilise") is not None else km_defaut
                    )
                    for t in jours_transport
                )

            ligne[f"Montant final — {libelle}"] = round(
                min(sum(montants), PLAFOND_MENSUEL), 2
            )

        # Les utilisateurs sans trajets sur les deux mois ne figurent pas dans l'export.
        if a_des_trajets:
            lignes.append(ligne)

    if not lignes:
        st.info("Aucun trajet à exporter pour ces deux mois.")
        st.stop()

    df_export = pd.DataFrame(lignes)
    colonnes_techniques = ["_modifie_prec", "_modifie_courant"]

    # Retire les colonnes de transport totalement vides (zéro pour tous les utilisateurs).
    colonnes_vides = [
        col for col in df_export.columns
        if col not in ["Utilisateur", *colonnes_techniques]
        and not col.startswith("Montant final")
        and (df_export[col] == 0).all()
    ]
    df_export = df_export.drop(columns=colonnes_vides)

    colonnes_montant = [
        col for col in df_export.columns if col.startswith("Montant final")
    ]

    def style_export(row):
        styles = ["" for _ in row.index]
        ligne_source = df_export.loc[row.name]
        for index, col in enumerate(row.index):
            if col == f"Montant final — {MOIS_FR[mois_prec_num - 1]} {annee_prec}" and ligne_source["_modifie_prec"]:
                styles[index] = "color: #d00000; font-weight: bold"
            if col == f"Montant final — {MOIS_FR[mois_num - 1]} {annee}" and ligne_source["_modifie_courant"]:
                styles[index] = "color: #d00000; font-weight: bold"
        return styles

    df_affichage = df_export.drop(columns=colonnes_techniques)
    st.dataframe(
        df_affichage.style.apply(style_export, axis=1).format(
            {col: "{:.2f} €" for col in colonnes_montant}
        ),
        hide_index=True,
        width="stretch"
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_affichage.to_excel(writer, index=False, sheet_name="Indemnités")
        ws = writer.sheets["Indemnités"]
        from openpyxl.styles import Font
        rouge_gras = Font(color="D00000", bold=True)

        for col_index, titre in enumerate(df_affichage.columns, start=1):
            ws.cell(1, col_index).font = Font(bold=True)
            ws.column_dimensions[ws.cell(1, col_index).column_letter].width = max(14, len(titre) + 2)

        for ligne_index, ligne in df_export.iterrows():
            if ligne["_modifie_prec"]:
                titre = f"Montant final — {MOIS_FR[mois_prec_num - 1]} {annee_prec}"
                ws.cell(ligne_index + 2, df_affichage.columns.get_loc(titre) + 1).font = rouge_gras
            if ligne["_modifie_courant"]:
                titre = f"Montant final — {MOIS_FR[mois_num - 1]} {annee}"
                ws.cell(ligne_index + 2, df_affichage.columns.get_loc(titre) + 1).font = rouge_gras

    buffer.seek(0)
    st.download_button(
        "Télécharger l'export Excel",
        data=buffer,
        file_name=f"Export_indemnites_{mois_courant}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if menu == "Validation":

    if not is_admin:
        st.error("Accès réservé aux administrateurs.")
        st.stop()

    st.header("✅ Validation des indemnités")
    st.divider()

    # ==================================================
    # VALIDATION GLOBALE
    # ==================================================
    st.subheader("Validation globale")

    if st.button("✅ Valider tous les utilisateurs"):

        res_users = (
            supabase
            .table("users")
            .select("id, km, nom, prenom")
            .neq("login", "admin")
            .order("nom")
            .execute()
        )

        users = res_users.data or []

        if not users:
            st.warning("Aucun utilisateur à valider.")
        else:

            succes = 0
            erreurs = []

            progress = st.progress(0)

            for i, u in enumerate(users):

                try:

                    valider_mois(
                        user_id=int(u["id"]),
                        annee=annee,
                        mois=mois_num,
                        km=float(u["km"] or 0)
                    )

                    succes += 1

                except Exception as e:

                    erreurs.append(
                        f"{u['prenom']} {u['nom']} : {e}"
                    )

                progress.progress((i + 1) / len(users))

            progress.empty()

            if succes:
                st.success(
                    f"{succes} utilisateur(s) validé(s)."
                )

            if erreurs:

                st.error(
                    "Certaines validations ont échoué :"
                )

                for err in erreurs:
                    st.write(err)

    st.divider()

    # ==================================================
    # VALIDATION INDIVIDUELLE
    # ==================================================
    st.subheader("Validation individuelle")

    res_users = (
        supabase
        .table("users")
        .select("id, prenom, nom, km")
        .neq("login", "admin")
        .order("nom")
        .execute()
    )

    users = res_users.data or []

    if not users:
        st.info("Aucun utilisateur.")
        st.stop()

    labels = [
        f"{u['prenom']} {u['nom']}"
        for u in users
    ]

    choix = st.selectbox(
        "Utilisateur",
        labels
    )

    user = next(
        u for u in users
        if f"{u['prenom']} {u['nom']}" == choix
    )

    user_id = int(user["id"])
    km_user = float(user["km"] or 0)

    mois_selectionne = f"{annee}-{mois_num:02d}"

    st.caption(
        f"📅 Validation du mois : {MOIS_FR[mois_num-1]} {annee}"
    )

    st.divider()
        # ==================================================
    # DÉTAIL DES TRAJETS
    # ==================================================

    start, end, _ = get_periode_reference(
        annee,
        mois_num
    )

    res_jours = (
        supabase
        .table("trajets")
        .select("jour, transport, km_utilise")
        .eq("user_id", user_id)
        .gte("jour", start.isoformat())
        .lte("jour", end.isoformat())
        .order("jour")
        .execute()
    )

    jours = res_jours.data or []

    if not jours:

        st.warning("Aucun trajet pour ce mois.")

        brut = 0
        plafonne = 0

    else:

        df = pd.DataFrame(jours)

        df["Jour"] = pd.to_datetime(df["jour"]).dt.strftime("%d/%m/%Y")

        df["KM"] = df["km_utilise"].fillna(km_user)

        df["Taux"] = df["transport"].map(TAUX)

        df["Indemnité"] = (
            df["KM"] *
            df["Taux"]
        ).round(2)

        st.dataframe(
            df[
                [
                    "Jour",
                    "transport",
                    "KM",
                    "Indemnité"
                ]
            ],
            width="stretch"
        )

        brut = round(
            df["Indemnité"].sum(),
            2
        )

        plafonne = min(
            brut,
            PLAFOND_MENSUEL
        )

    st.divider()

    # ==================================================
    # VALIDATION EXISTANTE
    # ==================================================

    ancienne_validation = (
        supabase
        .table("validations")
        .select("plafonne")
        .eq("user_id", user_id)
        .eq("mois", mois_selectionne)
        .execute()
        .data
    )

    deja_valide = len(ancienne_validation) > 0

    ancien_plafonne = (
        float(ancienne_validation[0]["plafonne"])
        if deja_valide
        else 0
    )

    correction = round(
        plafonne - ancien_plafonne,
        2
    )

    # ==================================================
    # RÉGULARISATION EXISTANTE
    # ==================================================

    reg_existante = (
        supabase
        .table("regularisations")
        .select("montant")
        .eq("user_id", user_id)
        .eq("mois_source", mois_selectionne)
        .execute()
        .data
    )

    ancienne_reg = (
        float(reg_existante[0]["montant"])
        if reg_existante
        else 0
    )

    st.subheader("Simulation")

    c1, c2, c3, c4 = st.columns(4)

    c1.metric(
        "Montant brut",
        f"{brut:.2f} €"
    )

    c2.metric(
        "Plafonné",
        f"{plafonne:.2f} €"
    )

    if deja_valide:

        c3.metric(
            "Ancienne validation",
            f"{ancien_plafonne:.2f} €"
        )

    else:

        c3.metric(
            "Validation",
            "Première"
        )

    if deja_valide:

        c4.metric(
            "Correction",
            f"{correction:+.2f} €"
        )

    else:

        c4.metric(
            "Correction",
            "0.00 €"
        )

    if deja_valide:

        st.info(
            f"""
Ancien montant validé : **{ancien_plafonne:.2f} €**

Nouveau montant : **{plafonne:.2f} €**

Régularisation actuelle : **{ancienne_reg:+.2f} €**

Nouvelle régularisation qui sera créée : **{correction:+.2f} €**
"""
        )

    else:

        st.success(
            f"""
Première validation.

Le montant payé sera de **{plafonne:.2f} €**.
"""
        )

    st.divider()
        # ==================================================
    # VALIDATION
    # ==================================================

    col1, col2 = st.columns(2)

    with col1:

        if st.button("✅ Valider le mois"):

            try:

                resultat = valider_mois(
                    user_id=user_id,
                    annee=annee,
                    mois=mois_num,
                    km=km_user
                )

                if not deja_valide:

                    st.success(
                        f"""
✅ Première validation enregistrée.

Montant brut : {resultat['brut']:.2f} €

Montant plafonné : {resultat['plafonne']:.2f} €
"""
                    )

                else:

                    if correction == 0:

                        st.success(
                            f"""
✅ Validation mise à jour.

Aucune modification du montant.

Montant payé : {resultat['plafonne']:.2f} €
"""
                        )

                    else:

                        if correction > 0:

                            texte = (
                                f"Une régularisation de "
                                f"+{correction:.2f} € "
                                f"sera ajoutée au mois suivant."
                            )

                        else:

                            texte = (
                                f"Une régularisation de "
                                f"{correction:.2f} € "
                                f"sera déduite du mois suivant."
                            )

                        st.success(
                            f"""
✅ Validation mise à jour.

Ancien montant :
{ancien_plafonne:.2f} €

Nouveau montant :
{plafonne:.2f} €

{texte}
"""
                        )

                st.rerun()

            except Exception as e:

                st.error(
                    f"Erreur : {e}"
                )

    with col2:

        if deja_valide:

            st.info(
                """
ℹ️ Ce mois a déjà été validé.

Toute modification des trajets entraînera automatiquement
une mise à jour de la régularisation du mois suivant.

Une seule régularisation est conservée pour ce mois.
                """
            )

        else:

            st.info(
                """
ℹ️ Première validation.

Aucune régularisation ne sera créée.
"""
            )





# Le traitement du menu Exports est volontairement placé après les autres sections.
if menu == "Exports":
    afficher_exports()

